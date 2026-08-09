#!/usr/bin/env python3
"""
Extract credit-risk supervisory validation rules to committed JSON.

Reads the two raw validation-rule workbooks in docs/assets/ (gitignored, fetched
by scripts/download_docs.py) and emits a filtered, machine-readable JSON extract
covering only the templates this project produces:

- EBA DPM 3.0(3.0.1) validation rules  -> COREP C 02.00 / C 07.00 / C 08.0x /
  C 09.0x / C 34.xx  (the CRR framework)
- BoE banking_reporting v4.0.0 rules   -> OF02 / OF07 / OF08 / OF09 / C08.04 /
  C09.04 / C34.xx    (the Basel 3.1 framework)

The workbooks are xlsx (gitignored); the JSON extracts are the committed
artefact, so downstream consumers and CI never need the raw files.

Usage:
    uv run python scripts/extract_validation_rules.py
    uv run python scripts/extract_validation_rules.py --check
    uv run python scripts/extract_validation_rules.py --sample 3

References:
- docs/reference/validation-rules/index.md — provenance, schema, formula grammar
- EBA validation rules: https://www.eba.europa.eu/risk-and-data-analysis/reporting-frameworks
- BoE banking taxonomy: https://www.bankofengland.co.uk/statistics/data-collection
"""

from __future__ import annotations

import argparse
import datetime as dt
import json
import re
import sys
from dataclasses import asdict, dataclass
from pathlib import Path
from typing import TYPE_CHECKING, Any, Literal

from openpyxl import load_workbook

if TYPE_CHECKING:
    from collections.abc import Callable

# Project root (parent of scripts directory)
PROJECT_ROOT = Path(__file__).parent.parent
ASSETS_DIR = PROJECT_ROOT / "docs" / "assets"
# The extracts are PACKAGE DATA, not documentation assets: the evaluator in
# rwa_calc.reporting.validations reads them at runtime through importlib.resources,
# so they must ship inside the wheel (pyproject packages only src/rwa_calc).
OUTPUT_DIR = PROJECT_ROOT / "src" / "rwa_calc" / "reporting" / "validations" / "rules"

RETRIEVED = "2026-08-01"

# ── EBA source (CRR framework) ───────────────────────────────────────────────

EBA_FILE = ASSETS_DIR / "eba-validation-rules.xlsx"
EBA_SHEET = "v3.0(3.0.1)"
EBA_VERSION = "3.0(3.0.1)"
EBA_URL = (
    "https://www.eba.europa.eu/sites/default/files/2026-06/"
    "12d2a6ae-9f58-47ab-a684-cdc9924ed4aa/"
    "%28up%20to%203.5%29%20EBA_validation_rules_2026-06-10.xlsx"
)
EBA_LANDING = "https://www.eba.europa.eu/risk-and-data-analysis/reporting/reporting-frameworks"
EBA_OUTPUT = OUTPUT_DIR / "crr-eba-v3.0-credit-risk.json"

# Table codes carry a space after "C" in this workbook (e.g. "C 07.00.a").
EBA_TABLE_PREFIXES = ("C 02.00", "C 07.00", "C 08.0", "C 09.0", "C 34.")

# 0-indexed column positions on the EBA sheet.
EBA_COL_ID = 0
EBA_COL_REPLACES = 1
EBA_COL_CHANGED_IN_RELEASE = 2
EBA_COL_LAST_CHANGE = 3
EBA_COL_DEACTIVATED = 4
EBA_COL_REACTIVATED = 5
EBA_COL_DELETED = 6
EBA_COL_NOT_IN_XBRL = 7
EBA_COL_TYPE = 8
EBA_COL_SEVERITY = 9
EBA_COL_TABLES = range(10, 17)  # T1..T7
EBA_COL_ROWS = 17
EBA_COL_COLUMNS = 18
EBA_COL_SHEETS = 19
EBA_COL_FORMULA = 20
EBA_COL_PREREQUISITES = 21
EBA_COL_IF_VALUE_MISSING = 22
EBA_COL_ARITHMETIC = 23
EBA_COL_NARRATIVE = 24

# ── BoE source (Basel 3.1 framework) ─────────────────────────────────────────

BOE_FILE = ASSETS_DIR / "boe-validation-rules-banking-reporting-v4.0.0.xlsx"
BOE_SHEET = "banking_reporting"
BOE_VERSION = "4.0.0"
BOE_URL = (
    "https://www.bankofengland.co.uk/-/media/boe/files/prudential-regulation/"
    "regulatory-reporting/banking/2026/february/boebankingtaxonomyvalidationsv400.zip"
)
BOE_LANDING = (
    "https://www.bankofengland.co.uk/prudential-regulation/regulatory-reporting/"
    "regulatory-reporting-banking-sector"
)
# The BoE workbook ships inside the taxonomy zip; download_docs.py extracts this member.
BOE_ARCHIVE_MEMBER = "Bank of England Banking Taxonomy Validations Banking reporting v4.0.0.xlsx"
BOE_OUTPUT = OUTPUT_DIR / "basel31-boe-v4.0.0-credit-risk.json"

# BoE table codes carry NO space after "C" (e.g. "C34.02.01.01", "OF07.00.01.01").
BOE_TABLE_PREFIXES = ("OF02", "OF07", "OF08", "OF09", "C08.04", "C09.04", "C34.")

# 0-indexed column positions on the BoE sheet.
BOE_COL_OWNER = 0
BOE_COL_FRAMEWORK = 1
BOE_COL_FRAMEWORK_VERSION = 2
BOE_COL_RULE_CODE = 3
BOE_COL_SCOPE = 4
BOE_COL_WHERE = 5
BOE_COL_JOIN = 6
BOE_COL_EXPRESSION = 7
BOE_COL_SIMPLE_EXPRESSION = 8
BOE_COL_PRECONDITION = 9
BOE_COL_SIMPLE_PRECONDITION = 10
BOE_COL_IN_XBRL = 11
BOE_COL_DEACTIVATED = 12
BOE_COL_ALWAYS_EXECUTE = 13
BOE_COL_ERROR_MESSAGE = 14
BOE_COL_LABEL = 15
BOE_COL_SEVERITY = 16
BOE_COL_TABLES = range(17, 21)  # T1..T4

# ── Parsing patterns ─────────────────────────────────────────────────────────

# "SHORT_LABEL(en) - some text|" / "DESCRIPTION(en) - v5745_q|"
#
# Two changes keep this linear. The body is `[^|]*` rather than a lazy `.*?`:
# excluding the terminator makes the scan to it deterministic, where `\s*` beside
# a dot-matches-all `.*?` is ambiguous over the separator's spaces. And the
# leading run is gated by `(?<![A-Z_])`, so a match can only START where the run
# does — without it the scan re-enters the same run at every offset, which is
# quadratic in a long unmatched run of capitals. A leftmost match always begins
# at the run's start anyway, so the gate rejects only positions that could never
# have matched. Leading space is left in and stripped by the caller.
LABEL_SEGMENT = re.compile(r"(?<![A-Z_])([A-Z_]+)\(([a-z]{2})\)\s*-([^|]*)\|")

# EBA rule identifiers as they appear inside BoE DESCRIPTION segments: v5745_q, e4893_n
EBA_RULE_ID = re.compile(r"\b([ve]\d+_[a-z]+)\b")

# "WARNING - PRA001|" -> severity token + module codes. The cell's single trailing
# "|" is removed by the caller before matching rather than by an optional `\|?`
# here: that trailing group sat between a lazy `(.*?)` and a `\s*$`, and the three
# together made every tail position a candidate split — quadratic on a cell that
# does not match at all.
BOE_SEVERITY = re.compile(r"^([A-Za-z]+)\s*(?:-(.*))?$", re.DOTALL)

# A scope id token is "well-formed" only at the current 4-digit DPM width.
FOUR_DIGIT_ID = re.compile(r"^\d{4}$")

# `(0010-0095)` / `(010-140)` — an inclusive span rather than a single id.
ID_RANGE = re.compile(r"^\d+-\d+$")

# Any digit run shorter than 4 belongs to the pre-2014 DPM numbering.
LEGACY_DIGIT_RUN = re.compile(r"\b\d{1,3}\b")


@dataclass(frozen=True)
class Scope:
    """A row / column / sheet scope drawn from a `(0010;0020)`-style cell.

    `kind` is "none" (empty cell), "all" (the literal `(All)`) or "list".
    `ids` holds the tokens verbatim — see the README on legacy 3-digit ids.
    """

    kind: Literal["none", "all", "list"]
    ids: tuple[str, ...]
    has_legacy_ids: bool
    has_id_ranges: bool


@dataclass(frozen=True)
class EbaRule:
    """One EBA validation rule scoped to a credit-risk COREP table."""

    id: str
    severity: str
    type: str
    status: tuple[str, ...]
    tables: tuple[str, ...]
    rows: tuple[str, ...]
    rows_scope: str
    columns: tuple[str, ...]
    columns_scope: str
    sheets: tuple[str, ...]
    sheets_scope: str
    formula: str | None
    prerequisites: str | None
    if_value_missing: str | None
    arithmetic_approach: str | None
    narrative: str | None
    replaces: str | None
    changed_in_release: str | None
    last_change: str | None
    deactivated_on: str | None
    reactivated_on: str | None
    has_legacy_ids: bool
    has_id_ranges: bool


@dataclass(frozen=True)
class BoeRule:
    """One BoE validation rule scoped to a credit-risk OF / C table."""

    id: str
    severity: str
    severity_modules: tuple[str, ...]
    status: tuple[str, ...]
    tables: tuple[str, ...]
    expression: str | None
    expression_raw: str | None
    precondition: str | None
    precondition_raw: str | None
    scope: str | None
    where: str | None
    join: str | None
    short_label: str | None
    description: str | None
    eba_equivalent: str | None
    eba_equivalents: tuple[str, ...]
    error_message: str | None
    labels: dict[str, str]
    include_in_xbrl: bool
    always_execute: bool
    owner: str | None
    framework_code: str | None
    framework_version_code: str | None


# ── Public API ───────────────────────────────────────────────────────────────


def main() -> int:
    """Entry point: parse arguments and run the extraction workflow."""
    parser = argparse.ArgumentParser(
        description="Extract credit-risk supervisory validation rules to committed JSON.",
    )
    parser.add_argument(
        "--check",
        action="store_true",
        help="re-extract and fail non-zero if the committed JSON would change (for CI)",
    )
    parser.add_argument(
        "--sample",
        type=int,
        default=0,
        metavar="N",
        help="print N parsed rules per source to stdout for inspection",
    )
    args = parser.parse_args()

    if not EBA_FILE.exists() or not BOE_FILE.exists():
        missing = [str(p) for p in (EBA_FILE, BOE_FILE) if not p.exists()]
        print("Missing source workbook(s):")
        for path in missing:
            print(f"  - {path}")
        print("\nRun: uv run python scripts/download_docs.py")
        return 1

    eba_payload = extract_eba()
    boe_payload = extract_boe()

    if args.sample:
        _print_samples(eba_payload, boe_payload, args.sample)

    outputs = ((EBA_OUTPUT, eba_payload), (BOE_OUTPUT, boe_payload))

    if args.check:
        return _check_outputs(outputs)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    for path, payload in outputs:
        path.write_text(_dump_json(payload), encoding="utf-8")
        counts = payload["filter"]
        print(
            f"  wrote {path.relative_to(PROJECT_ROOT).as_posix()} "
            f"({counts['matched']} rules, {counts['live']} live, "
            f"{_fmt_size(path.stat().st_size)})"
        )
    return 0


def extract_eba() -> dict[str, Any]:
    """Extract credit-risk rules from the EBA workbook into a JSON-ready payload."""
    rules = [_parse_eba_row(row) for row in _matching_rows(EBA_FILE, EBA_SHEET, _eba_tables)]
    live = sum(1 for r in rules if r.status == ("live",))
    reactivated = sum(1 for r in rules if r.reactivated_on and "deleted" not in r.status)
    return {
        "source": {
            "framework": "CRR",
            "publisher": "EBA",
            "file": EBA_FILE.name,
            "sheet": EBA_SHEET,
            "url": EBA_URL,
            "landing_page": EBA_LANDING,
            "retrieved": RETRIEVED,
            "framework_version": EBA_VERSION,
        },
        "filter": {
            "table_prefixes": list(EBA_TABLE_PREFIXES),
            "matched": len(rules),
            "live": live,
            "live_or_reactivated": live + reactivated,
        },
        "rules": [_rule_to_dict(r) for r in rules],
    }


def extract_boe() -> dict[str, Any]:
    """Extract credit-risk rules from the BoE workbook into a JSON-ready payload."""
    rules = [_parse_boe_row(row) for row in _matching_rows(BOE_FILE, BOE_SHEET, _boe_tables)]
    live = sum(1 for r in rules if r.status == ("live",))
    not_deactivated = sum(1 for r in rules if "deactivated" not in r.status)
    return {
        "source": {
            "framework": "Basel 3.1",
            "publisher": "BoE",
            "file": BOE_FILE.name,
            "sheet": BOE_SHEET,
            "url": BOE_URL,
            "archive_member": BOE_ARCHIVE_MEMBER,
            "landing_page": BOE_LANDING,
            "retrieved": RETRIEVED,
            "framework_version": BOE_VERSION,
        },
        "filter": {
            "table_prefixes": list(BOE_TABLE_PREFIXES),
            "matched": len(rules),
            "live": live,
            "not_deactivated": not_deactivated,
            "in_xbrl": sum(1 for r in rules if r.include_in_xbrl),
        },
        "rules": [_rule_to_dict(r) for r in rules],
    }


# ── Private helpers ──────────────────────────────────────────────────────────


def _matching_rows(
    path: Path,
    sheet: str,
    tables_of: Callable[[tuple[Any, ...]], tuple[str, ...]],
) -> list[tuple[Any, ...]]:
    """Return data rows whose table codes match the source's prefix filter."""
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        rows = workbook[sheet].iter_rows(values_only=True)
        next(rows)  # discard the header row
        return [row for row in rows if row and row[0] and tables_of(row)]
    finally:
        workbook.close()


def _eba_tables(row: tuple[Any, ...]) -> tuple[str, ...]:
    """Return the EBA table codes on a row that match the credit-risk prefixes."""
    codes = tuple(_clean(row[i]) for i in EBA_COL_TABLES if _clean(row[i]))
    return codes if any(c.startswith(EBA_TABLE_PREFIXES) for c in codes) else ()


def _boe_tables(row: tuple[Any, ...]) -> tuple[str, ...]:
    """Return the BoE table codes on a row that match the credit-risk prefixes."""
    codes = tuple(_clean(row[i]) for i in BOE_COL_TABLES if _clean(row[i]))
    return codes if any(c.startswith(BOE_TABLE_PREFIXES) for c in codes) else ()


def _parse_eba_row(row: tuple[Any, ...]) -> EbaRule:
    """Build an EbaRule from a raw worksheet row."""
    rows_scope = _parse_scope(row[EBA_COL_ROWS])
    cols_scope = _parse_scope(row[EBA_COL_COLUMNS])
    sheets_scope = _parse_scope(row[EBA_COL_SHEETS])
    return EbaRule(
        id=_clean(row[EBA_COL_ID]),
        severity=_clean(row[EBA_COL_SEVERITY]).upper(),
        type=_clean(row[EBA_COL_TYPE]),
        status=_eba_status(row),
        tables=_eba_tables(row),
        rows=rows_scope.ids,
        rows_scope=rows_scope.kind,
        columns=cols_scope.ids,
        columns_scope=cols_scope.kind,
        sheets=sheets_scope.ids,
        sheets_scope=sheets_scope.kind,
        formula=_clean_or_none(row[EBA_COL_FORMULA]),
        prerequisites=_clean_or_none(row[EBA_COL_PREREQUISITES]),
        if_value_missing=_clean_or_none(row[EBA_COL_IF_VALUE_MISSING]),
        arithmetic_approach=_clean_or_none(row[EBA_COL_ARITHMETIC]),
        narrative=_clean_or_none(row[EBA_COL_NARRATIVE]),
        replaces=_clean_or_none(row[EBA_COL_REPLACES]),
        changed_in_release=_clean_or_none(row[EBA_COL_CHANGED_IN_RELEASE]),
        last_change=_clean_or_none(row[EBA_COL_LAST_CHANGE]),
        deactivated_on=_as_date(row[EBA_COL_DEACTIVATED]),
        reactivated_on=_as_date(row[EBA_COL_REACTIVATED]),
        has_legacy_ids=any(s.has_legacy_ids for s in (rows_scope, cols_scope, sheets_scope)),
        has_id_ranges=any(s.has_id_ranges for s in (rows_scope, cols_scope, sheets_scope)),
    )


def _parse_boe_row(row: tuple[Any, ...]) -> BoeRule:
    """Build a BoeRule from a raw worksheet row."""
    severity, modules = _parse_boe_severity(row[BOE_COL_SEVERITY])
    labels = _parse_label_segments(row[BOE_COL_LABEL])
    description = labels.get("DESCRIPTION")
    equivalents = tuple(EBA_RULE_ID.findall(description)) if description else ()
    error_messages = _parse_label_segments(row[BOE_COL_ERROR_MESSAGE])
    return BoeRule(
        id=_clean(row[BOE_COL_RULE_CODE]),
        severity=severity,
        severity_modules=modules,
        status=_boe_status(row),
        tables=_boe_tables(row),
        expression=_clean_or_none(row[BOE_COL_SIMPLE_EXPRESSION]),
        expression_raw=_clean_or_none(row[BOE_COL_EXPRESSION]),
        precondition=_clean_or_none(row[BOE_COL_SIMPLE_PRECONDITION]),
        precondition_raw=_clean_or_none(row[BOE_COL_PRECONDITION]),
        scope=_clean_or_none(row[BOE_COL_SCOPE]),
        where=_clean_or_none(row[BOE_COL_WHERE]),
        join=_clean_or_none(row[BOE_COL_JOIN]),
        short_label=labels.get("SHORT_LABEL"),
        description=description,
        eba_equivalent=equivalents[0] if len(equivalents) == 1 else None,
        eba_equivalents=equivalents,
        error_message=error_messages.get("BUSINESS"),
        labels=labels,
        include_in_xbrl=_clean(row[BOE_COL_IN_XBRL]).lower() == "yes",
        always_execute=_clean(row[BOE_COL_ALWAYS_EXECUTE]).lower() == "yes",
        owner=_clean_or_none(row[BOE_COL_OWNER]),
        framework_code=_clean_or_none(row[BOE_COL_FRAMEWORK]),
        framework_version_code=_clean_or_none(row[BOE_COL_FRAMEWORK_VERSION]),
    )


def _eba_status(row: tuple[Any, ...]) -> tuple[str, ...]:
    """Derive the status flags for an EBA rule; ("live",) when none apply."""
    flags: list[str] = []
    if _clean(row[EBA_COL_DEACTIVATED]):
        flags.append("deactivated")
    if _clean(row[EBA_COL_DELETED]):
        flags.append("deleted")
    if _clean(row[EBA_COL_NOT_IN_XBRL]):
        flags.append("not_in_xbrl")
    return tuple(flags) if flags else ("live",)


def _boe_status(row: tuple[Any, ...]) -> tuple[str, ...]:
    """Derive the status flags for a BoE rule; ("live",) when none apply."""
    flags: list[str] = []
    if _clean(row[BOE_COL_DEACTIVATED]).lower() == "yes":
        flags.append("deactivated")
    if _clean(row[BOE_COL_IN_XBRL]).lower() == "no":
        flags.append("not_in_xbrl")
    return tuple(flags) if flags else ("live",)


def _parse_scope(value: Any) -> Scope:
    """Parse a `(0010;0020;0030)` / `(All)` / empty scope cell.

    Tokens are kept verbatim — sub-4-digit ids belong to the pre-2014 DPM
    numbering and are NOT zero-padded onto the modern grid. See the README.
    """
    text = _clean(value)
    if not text:
        return Scope(kind="none", ids=(), has_legacy_ids=False, has_id_ranges=False)

    inner = text.removeprefix("(").removesuffix(")").strip()
    if inner.lower() == "all":
        return Scope(kind="all", ids=(), has_legacy_ids=False, has_id_ranges=False)

    ids = tuple(token.strip() for token in inner.split(";") if token.strip())
    return Scope(
        kind="list",
        ids=ids,
        has_legacy_ids=any(
            not FOUR_DIGIT_ID.match(token) and LEGACY_DIGIT_RUN.search(token) for token in ids
        ),
        has_id_ranges=any(ID_RANGE.match(token) for token in ids),
    )


def _parse_boe_severity(value: Any) -> tuple[str, tuple[str, ...]]:
    """Split a `WARNING - PRA001|` cell into severity and module codes."""
    text = _clean(value)
    if not text:
        return "", ()
    match = BOE_SEVERITY.match(text.rstrip().removesuffix("|"))
    if not match:
        return text.upper(), ()
    severity = match.group(1).upper()
    modules = match.group(2) or ""
    return severity, tuple(m.strip() for m in modules.split(";") if m.strip())


def _parse_label_segments(value: Any) -> dict[str, str]:
    """Parse `KIND(en) - text|` segments into a {KIND: text} dict."""
    text = _clean(value)
    if not text:
        return {}
    return {kind: body.strip() for kind, _lang, body in LABEL_SEGMENT.findall(text)}


def _rule_to_dict(rule: EbaRule | BoeRule) -> dict[str, Any]:
    """Convert a rule dataclass to a JSON-ready dict (tuples become lists)."""
    return {key: list(val) if isinstance(val, tuple) else val for key, val in asdict(rule).items()}


def _dump_json(payload: dict[str, Any]) -> str:
    """Serialise a payload deterministically, with a trailing newline.

    `ensure_ascii=True` (the json default) is deliberate: the source labels carry
    typographic dashes and quotes, and a non-ASCII artefact would break any consumer
    doing `json.load(open(path))` on a cp1252 default locale. The committed files are
    pure ASCII and therefore encoding-agnostic.
    """
    return json.dumps(payload, indent=2, ensure_ascii=True) + "\n"


def _check_outputs(outputs: tuple[tuple[Path, dict[str, Any]], ...]) -> int:
    """Compare freshly extracted payloads against the committed JSON."""
    stale: list[str] = []
    for path, payload in outputs:
        rel = path.relative_to(PROJECT_ROOT).as_posix()
        if not path.exists():
            stale.append(f"{rel} (missing)")
            continue
        if path.read_text(encoding="utf-8") != _dump_json(payload):
            stale.append(f"{rel} (out of date)")
        else:
            print(f"  ok    {rel}")

    if stale:
        print("\nCommitted extract does not match the source workbooks:")
        for item in stale:
            print(f"  - {item}")
        print("\nRegenerate with: uv run python scripts/extract_validation_rules.py")
        return 1
    return 0


def _print_samples(eba: dict[str, Any], boe: dict[str, Any], count: int) -> None:
    """Print a handful of parsed rules from each source for eyeballing."""
    print(f"\nEBA sample ({eba['filter']['matched']} matched, {eba['filter']['live']} live)")
    print("-" * 78)
    for rule in eba["rules"][:count]:
        print(json.dumps(rule, indent=2))

    print(f"\nBoE sample ({boe['filter']['matched']} matched, {boe['filter']['live']} live)")
    print("-" * 78)
    for rule in boe["rules"][:count]:
        print(json.dumps(rule, indent=2))
    print()


def _clean(value: Any) -> str:
    """Normalise a worksheet cell to a stripped string ('' for blanks)."""
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.date().isoformat()
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value).strip()


def _clean_or_none(value: Any) -> str | None:
    """Like `_clean`, but blanks become None so they drop out of the JSON."""
    return _clean(value) or None


def _as_date(value: Any) -> str | None:
    """Return an ISO date string for a date cell, else the raw text, else None."""
    return _clean_or_none(value)


def _fmt_size(nbytes: int) -> str:
    """Format byte count as human-readable string."""
    if nbytes < 1024:
        return f"{nbytes} B"
    if nbytes < 1024 * 1024:
        return f"{nbytes / 1024:.1f} KB"
    return f"{nbytes / (1024 * 1024):.1f} MB"


if __name__ == "__main__":
    sys.exit(main())
